import json
import logging
import os
import pickle
import tempfile
import openpyxl
import pandas as pd
from collections import Counter
from copy import copy
from openpyxl import load_workbook
from openpyxl import Workbook
from typing import List, Tuple, Dict, Union


def open_file(path):
    """
    Opens [path] with a subprocess in order to continue running the main process.
    """
    import subprocess
    import platform
    import os

    if Path(path).exists():
        if platform.system() == 'Darwin':  # macOS
            subprocess.call(('open', path))
        elif platform.system() == 'Windows':  # Windows
            os.startfile(path)
        else:  # linux variants
            subprocess.call(('xdg-open', path))
    else:
        logging.getLogger().error(f"Error: {path} is not a file.")


def get_item_counts(lst):
    """
    Return list of tuples (item,item count in lst) ordered from most common to least.
    If lst is empty, returns empty list.
    :param lst:
    :return:
    """
    if len(lst) == 0:
        return []
    c = Counter(lst)
    return c.most_common(len(c))


def most_frequent_plus_frequency(lst):
    """
    Return (most frequent object in l, number of occurrences).
    If empty, returns None,0
    :param lst: list-like of objects
    :return: most frequent object in l, number of occurrences
    """
    occurrence_count = get_item_counts(lst)
    if len(occurrence_count) == 0:
        return None, 0
    return occurrence_count[0]


def tuple_list_to_string(list_of_tuples):
    """
    Return string with linebreaks between the tuples.
    :param list_of_tuples: list of tuples ( str, int )
    :return: 
    """
    return "\n".join([str(l) + ": " + str(r) for l, r in list_of_tuples])


def convert_lists_to_dataframe(list_of_column_data: List[list], columns: List[str]):
    """
    Convert list of columns to dataframe with columns [columns].
    :param list_of_column_data: List of columns. Each column is a list of values.
                                Assumes the length of all columns is the same.
    :param columns: List of column names
    :return: Pandas.DataFrame that represents the same data
    """
    return pd.DataFrame(data={col_name: values for col_name, values in zip(columns, list_of_column_data)})


def _auto_format_cell_dimensions_(ws, until_max_width=False):
    """
    Auto-fits the cell width and row height of worksheet [ws] in an already configured writer instance.
    Does not save.
    If until_max_width is True - auto fits cells' width until a max width, after which the text will wrap, and
    keeps the row height to a max of 120.
    If False, auto fits the column according to the column's longest cell content.
    :param ws: openpyxl.Worksheet instance
    :return:
    """
    from openpyxl.utils import get_column_letter

    MAX_WIDTH = 35
    for col_index, col in enumerate(ws.iter_cols()):
        maximum_value_for_col = 0
        for cell in col:
            val_to_check = max([len(line) for line in str(cell.value).split('\n')])
            if val_to_check > maximum_value_for_col:
                maximum_value_for_col = val_to_check
            if until_max_width:
                if maximum_value_for_col > MAX_WIDTH:
                    prev_alignment = copy(cell.alignment)
                    prev_alignment.wrap_text = True
                    cell.alignment = prev_alignment
                    maximum_value_for_col = MAX_WIDTH
                    continue
        ws.column_dimensions[get_column_letter(col_index + 1)].width = maximum_value_for_col + 2

    MAX_HEIGHT = 120
    for row_index, row in enumerate(ws.iter_rows(), 1):
        # Do not rely on ws.row_dimensions because if you append to an excel without opening the file, the row
        #  dimensions of the new rows are None. Hence, calculate the height yourself
        # 115.2 height is for 8 lines -> 1 row is 14.4 in height. The number of lines depends on the wrapping too
        if ws.row_dimensions[row_index].height is None:
            # calculate height for this row
            max_lines = 0
            for cell in row:
                lines = str(cell.value).split('\n')
                wrapped_lines = [line for line in lines if len(line) > MAX_WIDTH]
                num_lines = len(lines) + len(wrapped_lines)
                max_lines = max(max_lines, num_lines)
            height = 14.4 * max_lines
        else:
            height = ws.row_dimensions[row_index].height

        if height > MAX_HEIGHT:
            ws.row_dimensions[row_index].height = MAX_HEIGHT

    return


def autofit_dimensions_writer(excelwriter: pd.ExcelWriter, until_max_width=False):
    """
    Autofit column and row dimensions using the writer. Does not close.
    """
    try:
        book = excelwriter.workbook
    except AttributeError:
        book = excelwriter.book

    for ws in book:
        _auto_format_cell_dimensions_(ws, until_max_width)
        excelwriter.sheets[ws.title] = ws


def autofit_dimensions(excelfile, until_max_width=False):
    """
    Autofit each sheet's column width to the values.
    If until_max_width is True - auto fits cells' width until a max width, after which the text will wrap, and
    keeps the row height to a max of 120.
    If False, auto fits the column according to the column's longest cell content.
    :param until_max_width: True for auto-fitting until a max value
    :param excelfile: path to excel file or workbook instance
    :return:
    """
    writer = pd.ExcelWriter(excelfile, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    writer.book = openpyxl.load_workbook(excelfile)
    autofit_dimensions_writer(writer, until_max_width)

    writer.close()


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, startcol=None,
                       truncate_sheet=False, resize_columns=True, until_max_width=False,
                       na_rep='', **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file.
                       If False - appends the dataframe to the existing sheet.

      resize_columns: default = True . It resize all columns based on cell content width
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
      na_rep: default = 'NA'. If, instead of NaN, you want blank cells, just edit as follows: na_rep=''


    Returns: None

    *******************

    CONTRIBUTION:
    Current helper function generated by [Baggio]: https://stackoverflow.com/users/14302009/baggio?tab=profile
    Contributions to the current helper function: https://stackoverflow.com/users/4046632/buran?tab=profile
    Original helper function: (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)


    Features of the new helper function:
    1) Now it works with python 3.9 and latest versions of pd and openpxl
    ---> Fixed the error: "zipfile.BadZipFile: File is not a zip file".
    2) Now It resize all columns based on cell content width AND all variables will be visible (SEE "resize_columns")
    3) You can handle NaN,  if you want that NaN are displayed as NaN or as empty cells (SEE "na_rep")
    4) Added "startcol", you can decide to start to write from specific column, otherwise will start from col = 0

    *******************
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    try:
        f = open(filename)
        # Do something with the file
        f.close()
    except IOError:
        # file not accessible
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename)

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')

    try:
        # try to open an existing workbook
        writer.workbook = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.workbook.sheetnames:
            startrow = writer.workbook[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.workbook.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.workbook.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.workbook.remove(writer.workbook.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.workbook.create_sheet(sheet_name, idx)

        # copy existing sheets
        for ws in writer.workbook.worksheets:
            writer.sheets[ws.title] = ws

    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        # startrow = -1
        startrow = 0

    if startcol is None:
        startcol = 0

    # write out the new sheet
    try:
        df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, na_rep=na_rep, **to_excel_kwargs)
    except Exception as e:
        logging.getLogger().exception(e)

    if resize_columns:
        ws = writer.sheets[sheet_name]
        _auto_format_cell_dimensions_(ws, until_max_width)
    # save the workbook
    writer.close()


def upload_df_dict_to_json(df_dict: Dict[str, pd.DataFrame], path):
    """
    Export dataframe dictionary into JSON file located at [path]

    @param df_dict: dictionary from sheet name to dataframe.
    @param path: absolute path to JSON file.
    """
    import json

    class JSONEncoder(json.JSONEncoder):  # Recursive JSON encoder
        def default(self, obj):
            if hasattr(obj, 'to_json'):
                return obj.to_json(orient='columns')
            return super().default(obj)

    try:
        with open(path, 'w') as fp:
            json.dump(df_dict, fp, cls=JSONEncoder)
    except OverflowError as e:
        logging.getLogger().exception(f"Could not upload df_dict to json: {e}")


def json_to_df_dict(json_path):
    """Load JSON worksheet into dataframe dictionary"""
    import json
    json_d = json.load(open(json_path))
    df_dict = {}

    for sheet_name, data in json_d.items():
        if type(data) is str:
            df_dict[sheet_name] = pd.read_json(data, orient='columns').reset_index(drop=True)
        else:
            df_dict[sheet_name] = pd.DataFrame(data).reset_index(drop=True)

    return df_dict


def get_value(df: pd.DataFrame, search_col_name, search_val, get_col):
    """
    Return the value in column [get_col] in the first index in which [search_col_name] is equal [search_val]

    @param df: dataframe
    @param search_col_name: Column name to search
    @param search_val: Value to search in the column
    @param get_col: Column name of the wanted value
    """
    filtered_df = df.loc[df[search_col_name] == search_val]
    try:

        values = filtered_df[get_col].values
        return values[0]

    except IndexError as e:
        err = SearchValNotFound(search_col_name, search_val)
        logging.getLogger().error(err)
        raise err


def load_pickle(path):
    """
    Return dictionary loaded from pickle file at [path].

    returns None when an exception occurs.
    """
    try:
        with Path(path).open("rb") as fp:
            return pickle.load(fp)
    except Exception as e:
        logging.getLogger().critical(f"Exception: {e}")
        return None


def df_to_tmp_excel(df: pd.DataFrame, sheet_name: str):
    """
    Create a tmp excel file, dump df inside a sheet named sheet_name and open the file.
    """
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    df.to_excel(tmp.name, sheet_name=sheet_name, index=True, header=True)
    open_file(tmp.name)



